"""Extract email addresses (and optional per-row metadata) from a .xlsx file.

Two ways of pulling data out of a workbook are supported:

* ``extract_emails`` — walks every cell on every sheet and collects
  strings that look like an email address. Useful when the user pasted
  emails into an unformatted spreadsheet.
* ``inspect_workbook`` + ``extract_email_subject_pairs`` — for files
  that have a structured layout with one column for emails and another
  for per-recipient subjects, this lets the GUI ask the user which
  column is which and then read them as pairs.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

EMAIL_REGEX = re.compile(r"[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}")


def extract_emails(xlsx_path: str | Path) -> list[str]:
    """Return a de-duplicated list of email addresses found in the file.

    Order is preserved (first occurrence wins).
    """
    wb = load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)
    seen: set[str] = set()
    ordered: list[str] = []
    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                for value in row:
                    if not isinstance(value, str):
                        continue
                    for match in EMAIL_REGEX.findall(value):
                        normalized = match.strip().lower()
                        if normalized in seen:
                            continue
                        seen.add(normalized)
                        ordered.append(match.strip())
    finally:
        wb.close()
    return ordered


@dataclass
class SheetSummary:
    """Lightweight description of one worksheet inside a workbook.

    ``columns`` is one entry per column (A, B, C, …) in the order they
    appear in the sheet. Each entry includes the column letter, the
    detected header label (taken from row 1 when it isn't blank), and
    a short preview of values from row 2 onwards so the user can tell
    columns apart even when they aren't labelled.
    """

    name: str
    columns: list["ColumnSummary"] = field(default_factory=list)
    suggested_email_column: str | None = None


@dataclass
class ColumnSummary:
    letter: str
    header: str
    preview: list[str]
    has_email_like: bool


def inspect_workbook(xlsx_path: str | Path, max_preview_rows: int = 4) -> list[SheetSummary]:
    """Return one ``SheetSummary`` per worksheet in the file.

    For each column, the first non-empty value from row 1 becomes the
    header (or "(no header)" when blank), then up to ``max_preview_rows``
    later values are returned to help disambiguate columns. The first
    column that contains an email-like value is recorded as the
    suggested email column.
    """
    wb = load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)
    out: list[SheetSummary] = []
    try:
        for ws in wb.worksheets:
            max_col = ws.max_column or 0
            columns: list[ColumnSummary] = []
            suggested_email_col: str | None = None

            # Row 1 is treated as the header row.
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
            for col_index in range(1, max_col + 1):
                letter = get_column_letter(col_index)
                header_value = header_row[col_index - 1] if col_index - 1 < len(header_row) else None
                header = (
                    str(header_value).strip()
                    if header_value not in (None, "")
                    else "(no header)"
                )

                preview: list[str] = []
                has_email = False
                for row in ws.iter_rows(
                    min_row=2,
                    max_row=2 + max_preview_rows + 8,
                    min_col=col_index,
                    max_col=col_index,
                    values_only=True,
                ):
                    value = row[0] if row else None
                    if value in (None, ""):
                        continue
                    text = str(value).strip()
                    if EMAIL_REGEX.search(text):
                        has_email = True
                    if len(preview) < max_preview_rows:
                        preview.append(text)

                if has_email and suggested_email_col is None:
                    suggested_email_col = letter
                columns.append(
                    ColumnSummary(
                        letter=letter,
                        header=header,
                        preview=preview,
                        has_email_like=has_email,
                    )
                )
            out.append(
                SheetSummary(
                    name=ws.title,
                    columns=columns,
                    suggested_email_column=suggested_email_col,
                )
            )
    finally:
        wb.close()
    return out


def extract_email_subject_pairs(
    xlsx_path: str | Path,
    sheet_name: str,
    email_column: str,
    subject_column: str | None,
    skip_header_row: bool = True,
) -> list[tuple[str, str]]:
    """Read ``(email, subject)`` pairs from a specific sheet/columns.

    Rows where the email cell is missing or doesn't match the email
    regex are silently skipped. ``subject`` is an empty string when no
    subject column is specified or when the subject cell for the row
    is empty. Each email is normalised to lowercase for de-duplication
    but the original cased form is returned.
    """
    wb = load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)
    pairs: list[tuple[str, str]] = []
    seen: set[str] = set()
    try:
        ws = wb[sheet_name]
        from openpyxl.utils import column_index_from_string

        email_idx = column_index_from_string(email_column)
        subject_idx = column_index_from_string(subject_column) if subject_column else None
        first_row = 2 if skip_header_row else 1
        for row in ws.iter_rows(min_row=first_row, values_only=True):
            email_cell = row[email_idx - 1] if email_idx - 1 < len(row) else None
            if email_cell is None:
                continue
            email_match = EMAIL_REGEX.search(str(email_cell))
            if not email_match:
                continue
            email = email_match.group(0).strip()
            normalized = email.lower()
            if normalized in seen:
                continue
            seen.add(normalized)
            subject = ""
            if subject_idx is not None and subject_idx - 1 < len(row):
                subject_cell = row[subject_idx - 1]
                if subject_cell not in (None, ""):
                    subject = str(subject_cell).strip()
            pairs.append((email, subject))
    finally:
        wb.close()
    return pairs
